import csv
import io
import json
import logging
import re
import uuid
from collections import defaultdict
from datetime import date, datetime
from typing import Any
from urllib.error import HTTPError
from urllib.parse import parse_qs, urlencode, urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.entities import Organization, Person, Relationship, RelationshipEdge
from app.schemas.content_asset import ImportMapRequest, ImportMapResponse, ImportUploadResponse

logger = logging.getLogger(__name__)

_COLUMN_HINTS: dict[str, str] = {
    "full name": "person.full_name",
    "name": "person.full_name",
    "first name": "person.first_name",
    "last name": "person.last_name",
    "email": "person.email",
    "phone": "person.phone",
    "mobile": "person.phone",
    "number": "person.phone",
    "cell": "person.phone",
    "company": "organization.name",
    "organization": "organization.name",
    "title": "person.primary_role",
    "role": "person.primary_role",
    "main asset class": "person.tags",
    "asset class": "person.tags",
    "secondary role": "person.secondary_roles",
    "secondary roles": "person.secondary_roles",
    "source": "person.source",
    "stage": "person.relationship_stage",
    "relationship stage": "person.relationship_stage",
    "status": "person.relationship_stage",
    "notes": "person.notes_summary",
    "note": "person.notes_summary",
    "what i'm looking for": "person.notes_summary",
    "looking for": "person.notes_summary",
    "tags": "person.tags",
    "tag": "person.tags",
    "request type": "relationship.type",
    "target location": "metadata.raw",
    "target locations": "metadata.raw",
    "telegram": "metadata.raw",
    "telegram id": "metadata.raw",
    "lifetime value": "person.lifetime_value",
    "ltv": "person.lifetime_value",
    "referral value": "person.referral_value",
    "relationship strength": "person.relationship_strength_score",
    "strength": "person.relationship_strength_score",
    "website": "organization.website",
    "location": "organization.location",
    "city": "organization.location",
    "owner": "relationship.owner_user_id",
    "owner user": "relationship.owner_user_id",
    "relationship type": "relationship.type",
    "contact type": "relationship.type",
    "referred by email": "relationship.referrer_email",
    "referrer email": "relationship.referrer_email",
    "introduced by email": "relationship.referrer_email",
    "referred by": "relationship.referrer_name",
    "referrer": "relationship.referrer_name",
    "introduced by": "relationship.referrer_name",
    "parent contact email": "relationship.parent_email",
    "manager email": "relationship.parent_email",
    "parent contact": "relationship.parent_name",
    "manager": "relationship.parent_name",
    "linkedin": "metadata.raw",
    "instagram": "metadata.raw",
    "facebook": "metadata.raw",
    "address": "metadata.raw",
    "zip": "metadata.raw",
    "postal": "metadata.raw",
    "deal type": "metadata.raw",
    "amount": "metadata.raw",
    "close date": "metadata.raw",
}

_SOURCE_TYPE_TABLE_MAP: dict[str, str] = {
    "contacts": "people + relationships + organizations",
    "linkedin": "people + relationships + organizations",
    "webinar_attendees": "people + organizations",
    "story_viewers": "people + engagement_events",
    "podcast_leads": "people + organizations",
    "deal_list": "people + organizations",
    "vendor_list": "people + organizations",
    "buyer_leads": "people + organizations",
    "seller_leads": "people + organizations",
}

_ALLOWED_AI_FIELDS = sorted(set(_COLUMN_HINTS.values()))
_BATCH_SIZE = 500


def _normalize_header(value: Any) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower())
    return re.sub(r"\s+", " ", cleaned).strip()


def _bounded_text(value: Any, max_len: int) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    return text[:max_len]


def _normalize_email(value: Any) -> str | None:
    email = str(value or "").strip().lower()
    return email or None


def _normalize_phone(value: Any) -> str | None:
    digits = re.sub(r"\D+", "", str(value or ""))
    return digits or None


def _value_present(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _serialize_cell(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _split_full_name(full_name: str | None) -> tuple[str, str]:
    text = (full_name or "").strip()
    if not text:
        return "", ""
    parts = text.split(None, 1)
    return parts[0], parts[1] if len(parts) > 1 else ""


def _parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_tags(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    raw = str(value or "").strip()
    if not raw:
        return {}
    if raw.startswith("{") and raw.endswith("}"):
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                return parsed
        except Exception:
            pass
    tags = [item.strip() for item in re.split(r"[,;|]", raw) if item.strip()]
    return {"labels": tags}


def _parse_roles(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    raw = str(value or "").strip()
    if not raw:
        return []
    return [item.strip() for item in re.split(r"[,;|]", raw) if item.strip()]


def _first_value(row: dict[str, Any], headers: list[str]) -> Any:
    for header in headers:
        value = row.get(header)
        if _value_present(value):
            return value
    return None


def _build_target_map(mapping: dict[str, str]) -> dict[str, list[str]]:
    target_map: dict[str, list[str]] = defaultdict(list)
    for column, target in mapping.items():
        target_map[target].append(column)
    return dict(target_map)


def _make_unique_headers(values: list[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = defaultdict(int)
    for index, value in enumerate(values, start=1):
        base = str(value or "").strip() or f"Column {index}"
        seen[base] += 1
        headers.append(base if seen[base] == 1 else f"{base} ({seen[base]})")
    return headers


def _score_header_candidate(values: list[Any]) -> float:
    normalized = [_normalize_header(value) for value in values if _value_present(value)]
    if not normalized:
        return -1.0

    hint_matches = 0
    short_label_count = 0
    narrative_penalty = 0.0
    for item in normalized:
        if any(hint in item or item in hint for hint in _COLUMN_HINTS):
            hint_matches += 1
        if len(item) <= 32:
            short_label_count += 1
        if len(item) > 48:
            narrative_penalty += 1.5

    unique_count = len(set(normalized))
    if len(normalized) == 1 and len(normalized[0]) > 40:
        return -2.0

    return hint_matches * 3.0 + short_label_count * 0.35 + unique_count * 0.2 - narrative_penalty


def _detect_header_row(raw_rows: list[list[Any]]) -> int:
    search_window = raw_rows[:20]
    best_index = 0
    best_score = float("-inf")
    for index, row in enumerate(search_window):
        score = _score_header_candidate(row)
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def _rows_to_records(
    raw_rows: list[list[Any]],
    header_row: int | None = None,
) -> tuple[int, list[str], list[dict[str, Any]]]:
    if not raw_rows:
        return 0, [], []

    if header_row is not None:
        header_index = header_row - 1
        if header_index < 0 or header_index >= len(raw_rows):
            raise ValueError(f"Header row {header_row} is out of range for this sheet")
    else:
        header_index = _detect_header_row(raw_rows)
    headers = _make_unique_headers(raw_rows[header_index])
    width = max(len(headers), max(len(row) for row in raw_rows[header_index:]))
    if len(headers) < width:
        headers.extend([f"Column {index}" for index in range(len(headers) + 1, width + 1)])

    rows: list[dict[str, Any]] = []
    for raw_row in raw_rows[header_index + 1:]:
        padded = list(raw_row) + [None] * (len(headers) - len(raw_row))
        normalized = {header: _serialize_cell(value) for header, value in zip(headers, padded)}
        if any(_value_present(value) for value in normalized.values()):
            rows.append(normalized)
    return header_index, headers, rows


def _read_uploaded_rows(
    file_name: str,
    file_bytes: bytes,
    sheet_name: str | None,
    header_row: int | None = None,
) -> tuple[str | None, int | None, list[str], list[dict[str, Any]]]:
    suffix = file_name.lower().rsplit(".", 1)[-1] if "." in file_name else ""
    if suffix == "csv":
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        raw_rows = [list(row) for row in reader]
        header_index, headers, rows = _rows_to_records(raw_rows, header_row=header_row)
        return None, header_index + 1, headers, rows

    if suffix not in {"xlsx", "xlsm"}:
        raise ValueError("Unsupported file format. Please upload .xlsx, .xlsm, or .csv")

    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
    raw_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    if not raw_rows:
        return worksheet.title, None, [], []

    header_index, headers, rows = _rows_to_records(raw_rows, header_row=header_row)
    return worksheet.title, header_index + 1, headers, rows


def _chunked(values: list[str], size: int = 500) -> list[list[str]]:
    return [values[index:index + size] for index in range(0, len(values), size)]


def _extract_google_sheet_id(sheet_url: str) -> str | None:
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", sheet_url)
    return match.group(1) if match else None


def _extract_gid(sheet_url: str) -> str | None:
    parsed = urlparse(sheet_url)
    query_gid = parse_qs(parsed.query).get("gid", [None])[0]
    if query_gid:
        return query_gid
    fragment_match = re.search(r"gid=(\d+)", parsed.fragment or "")
    if fragment_match:
        return fragment_match.group(1)
    return None


def _build_google_export_request(sheet_url: str, sheet_name: str | None) -> tuple[str, str]:
    sheet_id = _extract_google_sheet_id(sheet_url)
    if not sheet_id:
        raise ValueError("Invalid Google Sheets URL. Expected a docs.google.com/spreadsheets link.")

    gid = _extract_gid(sheet_url)
    if gid and not sheet_name:
        export_url = (
            f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?"
            + urlencode({"format": "csv", "gid": gid})
        )
        return export_url, f"google-sheet-{sheet_id}-{gid}.csv"

    export_url = (
        f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?"
        + urlencode({"format": "xlsx"})
    )
    return export_url, f"google-sheet-{sheet_id}.xlsx"


def _download_google_sheet(sheet_url: str, sheet_name: str | None) -> tuple[str, bytes]:
    export_url, file_name = _build_google_export_request(sheet_url, sheet_name)
    request = Request(
        export_url,
        headers={
            "User-Agent": "RelateOS Importer/1.0",
            "Accept": "text/csv,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;q=0.9,*/*;q=0.8",
        },
    )
    try:
        with urlopen(request, timeout=60) as response:
            content = response.read()
            if not content:
                raise ValueError("Google Sheets export returned an empty file")
            return file_name, content
    except HTTPError as exc:
        if exc.code in {401, 403}:
            raise ValueError(
                "Google denied access to this sheet. Set sharing to 'Anyone with the link can view' or publish the sheet to the web, then retry."
            ) from exc
        if exc.code == 404:
            raise ValueError(
                "Google Sheet not found. Confirm the sheet URL and tab (gid) are correct."
            ) from exc
        raise ValueError(
            f"Could not fetch Google Sheet (HTTP {exc.code}). Make sure the sheet is publicly accessible."
        ) from exc
    except Exception as exc:
        raise ValueError(
            "Could not fetch Google Sheet. Make sure the sheet is shared publicly or published, and the URL is valid."
        ) from exc


class ImportService:
    @staticmethod
    def _ai_map_columns(payload: ImportMapRequest, unmapped_fields: list[str]) -> dict[str, str]:
        if not settings.openai_api_key or not unmapped_fields:
            return {}
        try:
            from openai import OpenAI

            client = OpenAI(api_key=settings.openai_api_key)
            prompt = (
                f"Source type: {payload.source_type}\n"
                f"Raw columns: {payload.raw_columns}\n"
                f"Sample rows: {payload.sample_rows[:5]}\n\n"
                "Map only the unmapped columns to one of these canonical fields:\n"
                f"{_ALLOWED_AI_FIELDS + ['metadata.raw']}\n\n"
                "Return a JSON object where each key is an unmapped column and each value is a canonical field name. "
                "If a column should simply be preserved without a direct structured field, map it to metadata.raw. "
                "Return only valid JSON."
            )
            raw = client.chat.completions.create(
                model=settings.openai_model or "gpt-4o",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1,
                response_format={"type": "json_object"},
            ).choices[0].message.content or "{}"
            parsed = json.loads(raw)
            results: dict[str, str] = {}
            for column, target in parsed.items():
                canonical = str(target)
                if column in unmapped_fields and canonical in _ALLOWED_AI_FIELDS + ["metadata.raw"]:
                    results[column] = canonical
            return results
        except Exception as exc:
            logger.warning("AI import mapping failed, using heuristic mapping only: %s", exc)
            return {}

    @staticmethod
    def map_import(payload: ImportMapRequest) -> ImportMapResponse:
        mapping: dict[str, str] = {}
        unmapped: list[str] = []
        warnings: list[str] = []

        for column in payload.raw_columns:
            normalized = _normalize_header(column)
            target = None
            for hint, candidate in _COLUMN_HINTS.items():
                if hint in normalized:
                    target = candidate
                    break
            if target:
                mapping[column] = target
            else:
                unmapped.append(column)

        ai_mapping = ImportService._ai_map_columns(payload, unmapped)
        for column, target in ai_mapping.items():
            mapping[column] = target
        unmapped = [column for column in unmapped if column not in ai_mapping]

        if unmapped:
            warnings.append(f"Preserving {len(unmapped)} unmapped column(s) in metadata: {', '.join(unmapped[:10])}")
            if len(unmapped) > 10:
                warnings.append(f"Additional unmapped columns preserved: {len(unmapped) - 10}")

        confidence = 1.0 - (len(unmapped) / max(len(payload.raw_columns), 1)) * 0.35
        if ai_mapping:
            confidence = min(1.0, confidence + 0.1)

        return ImportMapResponse(
            suggested_table=_SOURCE_TYPE_TABLE_MAP.get(payload.source_type, "people + relationships + organizations"),
            suggested_column_mapping=mapping,
            confidence=round(max(confidence, 0.4), 2),
            warnings=warnings,
            unmapped_fields=unmapped,
        )

    @staticmethod
    def import_contacts_file(
        db: Session,
        *,
        file_name: str,
        file_bytes: bytes,
        source_type: str,
        sheet_name: str | None = None,
        header_row: int | None = None,
    ) -> ImportUploadResponse:
        resolved_sheet_name, header_row_used, headers, rows = _read_uploaded_rows(
            file_name,
            file_bytes,
            sheet_name,
            header_row=header_row,
        )
        mapping_response = ImportService.map_import(
            ImportMapRequest(source_type=source_type, raw_columns=headers, sample_rows=rows[:10])
        )
        mapping = mapping_response.suggested_column_mapping
        target_map = _build_target_map(mapping)

        warnings = list(mapping_response.warnings)
        stats = {
            "rows_processed": 0,
            "rows_skipped": 0,
            "contacts_created": 0,
            "contacts_updated": 0,
            "organizations_created": 0,
            "relationships_created": 0,
            "relationship_edges_created": 0,
        }
        stored_extra_fields: set[str] = set(mapping_response.unmapped_fields)

        organizations = {
            (org.name or "").strip().lower(): org
            for org in db.query(Organization).all()
            if (org.name or "").strip()
        }

        emails = {
            _normalize_email(_first_value(row, target_map.get("person.email", [])))
            for row in rows
            if _normalize_email(_first_value(row, target_map.get("person.email", [])))
        }
        phones = {
            _normalize_phone(_first_value(row, target_map.get("person.phone", [])))
            for row in rows
            if _normalize_phone(_first_value(row, target_map.get("person.phone", [])))
        }
        referrer_emails = {
            _normalize_email(_first_value(row, target_map.get("relationship.referrer_email", [])))
            for row in rows
            if _normalize_email(_first_value(row, target_map.get("relationship.referrer_email", [])))
        }
        parent_emails = {
            _normalize_email(_first_value(row, target_map.get("relationship.parent_email", [])))
            for row in rows
            if _normalize_email(_first_value(row, target_map.get("relationship.parent_email", [])))
        }

        people_by_email: dict[str, Person] = {}
        people_by_phone: dict[str, Person] = {}
        people_by_name: dict[str, Person] = {}

        for chunk in _chunked(sorted(list(emails | referrer_emails | parent_emails))):
            if not chunk:
                continue
            for person in db.query(Person).filter(func.lower(Person.email).in_(chunk)).all():
                if person.email:
                    people_by_email[person.email.lower()] = person

        for chunk in _chunked(sorted(list(phones))):
            if not chunk:
                continue
            for person in db.query(Person).filter(Person.phone.in_(chunk)).all():
                normalized_phone = _normalize_phone(person.phone)
                if normalized_phone:
                    people_by_phone[normalized_phone] = person

        for person in db.query(Person).filter(or_(Person.email.isnot(None), Person.phone.isnot(None))).all():
            name_key = f"{(person.first_name or '').strip().lower()}|{(person.last_name or '').strip().lower()}|{str(person.organization_id or '')}"
            if name_key not in people_by_name:
                people_by_name[name_key] = person

        relationship_cache = {
            relationship.person_id: relationship
            for relationship in db.query(Relationship).all()
        }
        existing_edge_keys = {
            (str(edge.source_contact_id), str(edge.target_contact_id), edge.relationship_type)
            for edge in db.query(RelationshipEdge).all()
        }

        def get_or_create_organization(name: str | None, row: dict[str, Any]) -> Organization | None:
            normalized = (name or "").strip().lower()
            if not normalized:
                return None
            organization = organizations.get(normalized)
            if organization:
                return organization

            organization = Organization(
                id=uuid.uuid4(),
                name=(name or "").strip(),
                website=str(_first_value(row, target_map.get("organization.website", [])) or "").strip() or None,
                location=str(_first_value(row, target_map.get("organization.location", [])) or "").strip() or None,
                org_type="other",
            )
            db.add(organization)
            db.flush()
            organizations[normalized] = organization
            stats["organizations_created"] += 1
            return organization

        def ensure_relationship(person: Person, row: dict[str, Any]) -> None:
            relationship = relationship_cache.get(person.id)
            relationship_type_raw = _first_value(row, target_map.get("relationship.type", [])) or person.primary_role or "lead"
            relationship_type = _bounded_text(relationship_type_raw, 50) or "lead"
            if len(str(relationship_type_raw or "").strip()) > 50 and len(warnings) < 10:
                warnings.append(
                    f"Row relationship type exceeded 50 chars and was truncated: {str(relationship_type_raw).strip()[:80]}"
                )

            lifecycle_stage_raw = person.relationship_stage or "new"
            lifecycle_stage = _bounded_text(lifecycle_stage_raw, 50) or "new"
            strength = person.relationship_strength_score or 0.0
            owner_user_id_raw = _first_value(row, target_map.get("relationship.owner_user_id", []))
            owner_user_id = _bounded_text(owner_user_id_raw, 100)
            if not relationship:
                relationship = Relationship(
                    id=uuid.uuid4(),
                    person_id=person.id,
                    type=relationship_type,
                    lifecycle_stage=lifecycle_stage,
                    relationship_strength=strength,
                    owner_user_id=owner_user_id,
                )
                db.add(relationship)
                relationship_cache[person.id] = relationship
                stats["relationships_created"] += 1
                return
            relationship.type = relationship_type or relationship.type
            relationship.lifecycle_stage = lifecycle_stage or relationship.lifecycle_stage
            relationship.relationship_strength = strength or relationship.relationship_strength
            if owner_user_id:
                relationship.owner_user_id = owner_user_id

        def resolve_related_person(email: str | None, name: str | None, relationship_kind: str) -> Person | None:
            normalized_email = _normalize_email(email)
            if normalized_email and normalized_email in people_by_email:
                return people_by_email[normalized_email]

            first_name, last_name = _split_full_name(name)
            if not first_name and not normalized_email:
                return None

            placeholder = Person(
                id=uuid.uuid4(),
                first_name=first_name or "Unknown",
                last_name=last_name,
                email=normalized_email,
                source="import",
                notes_summary=f"Auto-created from {relationship_kind} column during import.",
            )
            db.add(placeholder)
            db.flush()
            if normalized_email:
                people_by_email[normalized_email] = placeholder
            name_key = f"{placeholder.first_name.strip().lower()}|{placeholder.last_name.strip().lower()}|"
            people_by_name[name_key] = placeholder
            stats["contacts_created"] += 1
            ensure_relationship(placeholder, {})
            return placeholder

        for index, row in enumerate(rows, start=1):
            full_name = str(_first_value(row, target_map.get("person.full_name", [])) or "").strip()
            first_name = str(_first_value(row, target_map.get("person.first_name", [])) or "").strip()
            last_name = str(_first_value(row, target_map.get("person.last_name", [])) or "").strip()
            if full_name and not first_name:
                first_name, last_name = _split_full_name(full_name)

            email = _normalize_email(_first_value(row, target_map.get("person.email", [])))
            phone = _normalize_phone(_first_value(row, target_map.get("person.phone", [])))
            organization_name = str(_first_value(row, target_map.get("organization.name", [])) or "").strip()
            organization = get_or_create_organization(organization_name, row)

            if not first_name and not last_name and not email and not phone:
                stats["rows_skipped"] += 1
                if len(warnings) < 10:
                    warnings.append(f"Skipped row {index + 1}: missing name, email, and phone")
                continue

            person = None
            if email:
                person = people_by_email.get(email)
            if not person and phone:
                person = people_by_phone.get(phone)

            name_org_key = f"{first_name.lower()}|{last_name.lower()}|{str(organization.id if organization else '')}"
            if not person and first_name:
                person = people_by_name.get(name_org_key)

            created = False
            if not person:
                person = Person(
                    id=uuid.uuid4(),
                    first_name=first_name or "Unknown",
                    last_name=last_name,
                    email=email,
                    phone=phone,
                    source=source_type,
                    organization_id=organization.id if organization else None,
                )
                db.add(person)
                db.flush()
                created = True
                stats["contacts_created"] += 1

            updated = created
            if first_name and person.first_name != first_name:
                person.first_name = first_name
                updated = True
            if last_name and person.last_name != last_name:
                person.last_name = last_name
                updated = True
            if email and person.email != email:
                person.email = email
                updated = True
            if phone and person.phone != phone:
                person.phone = phone
                updated = True
            if organization and person.organization_id != organization.id:
                person.organization_id = organization.id
                updated = True

            role_raw = _first_value(row, target_map.get("person.primary_role", []))
            role = _bounded_text(role_raw, 50)
            if len(str(role_raw or "").strip()) > 50 and len(warnings) < 10:
                warnings.append(
                    f"Primary role exceeded 50 chars and was truncated: {str(role_raw).strip()[:80]}"
                )
            secondary_roles = _parse_roles(_first_value(row, target_map.get("person.secondary_roles", [])))
            relationship_stage_raw = _first_value(row, target_map.get("person.relationship_stage", []))
            relationship_stage = _bounded_text(relationship_stage_raw, 50)
            source_raw = _first_value(row, target_map.get("person.source", [])) or source_type
            source = _bounded_text(source_raw, 50) or source_type
            notes_summary = str(_first_value(row, target_map.get("person.notes_summary", [])) or "").strip() or None
            tags = _parse_tags(_first_value(row, target_map.get("person.tags", [])))
            lifetime_value = _parse_float(_first_value(row, target_map.get("person.lifetime_value", [])))
            referral_value = _parse_float(_first_value(row, target_map.get("person.referral_value", [])))
            relationship_strength = _parse_float(_first_value(row, target_map.get("person.relationship_strength_score", [])))

            if role and person.primary_role != role:
                person.primary_role = role
                updated = True
            if secondary_roles and person.secondary_roles != secondary_roles:
                person.secondary_roles = secondary_roles
                updated = True
            if relationship_stage and person.relationship_stage != relationship_stage:
                person.relationship_stage = relationship_stage
                updated = True
            if source and person.source != source:
                person.source = source
                updated = True
            if notes_summary and person.notes_summary != notes_summary:
                person.notes_summary = notes_summary
                updated = True
            if tags:
                merged_tags = dict(person.tags or {})
                merged_tags.update(tags)
                if merged_tags != person.tags:
                    person.tags = merged_tags
                    updated = True
            if lifetime_value is not None and person.lifetime_value != lifetime_value:
                person.lifetime_value = lifetime_value
                updated = True
            if referral_value is not None and person.referral_value != referral_value:
                person.referral_value = referral_value
                updated = True
            if relationship_strength is not None and person.relationship_strength_score != relationship_strength:
                person.relationship_strength_score = relationship_strength
                updated = True

            extras: dict[str, Any] = {}
            for header, value in row.items():
                if not _value_present(value):
                    continue
                mapped_target = mapping.get(header)
                if mapped_target is None or mapped_target == "metadata.raw":
                    extras[header] = value
                elif mapped_target.startswith("metadata."):
                    extras[header] = value
            if extras:
                metadata = dict(person.metadata_json or {})
                import_fields = dict(metadata.get("import_fields") or {})
                import_fields.update({key: _serialize_cell(value) for key, value in extras.items()})
                metadata["import_fields"] = import_fields
                metadata["last_import_source"] = source_type
                metadata["last_import_file"] = file_name
                person.metadata_json = metadata
                stored_extra_fields.update(extras.keys())
                updated = True

            if created:
                pass
            elif updated:
                stats["contacts_updated"] += 1

            if email:
                people_by_email[email] = person
            if phone:
                people_by_phone[phone] = person
            people_by_name[name_org_key] = person

            parent_email = _normalize_email(_first_value(row, target_map.get("relationship.parent_email", [])))
            parent_name = str(_first_value(row, target_map.get("relationship.parent_name", [])) or "").strip() or None
            parent_person = resolve_related_person(parent_email, parent_name, "parent relationship")
            if parent_person and person.parent_contact_id != parent_person.id:
                person.parent_contact_id = parent_person.id
                updated = True

            ensure_relationship(person, row)

            referrer_email = _normalize_email(_first_value(row, target_map.get("relationship.referrer_email", [])))
            referrer_name = str(_first_value(row, target_map.get("relationship.referrer_name", [])) or "").strip() or None
            referrer_person = resolve_related_person(referrer_email, referrer_name, "referrer relationship")

            edge_specs = []
            if referrer_person and referrer_person.id != person.id:
                edge_specs.append((referrer_person, person, str(_first_value(row, target_map.get("relationship.type", [])) or "introduced_by")))
            if parent_person and parent_person.id != person.id:
                edge_specs.append((parent_person, person, "reports_to"))

            for source_person, target_person, relationship_type in edge_specs:
                edge_key = (str(source_person.id), str(target_person.id), relationship_type)
                if edge_key in existing_edge_keys:
                    continue
                db.add(
                    RelationshipEdge(
                        id=uuid.uuid4(),
                        source_contact_id=source_person.id,
                        target_contact_id=target_person.id,
                        organization_id=organization.id if organization else None,
                        relationship_type=relationship_type,
                        strength=relationship_strength or 1.0,
                        evidence={"source": "excel_import", "file_name": file_name},
                    )
                )
                existing_edge_keys.add(edge_key)
                stats["relationship_edges_created"] += 1

            stats["rows_processed"] += 1
            if index % _BATCH_SIZE == 0:
                db.commit()

        db.commit()
        return ImportUploadResponse(
            file_name=file_name,
            source_type=source_type,
            sheet_name=resolved_sheet_name,
            header_row_used=header_row_used,
            rows_processed=stats["rows_processed"],
            rows_skipped=stats["rows_skipped"],
            contacts_created=stats["contacts_created"],
            contacts_updated=stats["contacts_updated"],
            organizations_created=stats["organizations_created"],
            relationships_created=stats["relationships_created"],
            relationship_edges_created=stats["relationship_edges_created"],
            suggested_column_mapping=mapping,
            unmapped_columns=mapping_response.unmapped_fields,
            stored_extra_fields=sorted(stored_extra_fields),
            warnings=warnings,
        )

    @staticmethod
    def import_contacts_from_url(
        db: Session,
        *,
        sheet_url: str,
        source_type: str,
        sheet_name: str | None = None,
        header_row: int | None = None,
    ) -> ImportUploadResponse:
        normalized_url = str(sheet_url or "").strip()
        if not normalized_url:
            raise ValueError("Google Sheets URL is required")
        if "docs.google.com/spreadsheets/" not in normalized_url:
            raise ValueError("Only public Google Sheets URLs are supported right now")

        file_name, payload = _download_google_sheet(normalized_url, sheet_name)
        result = ImportService.import_contacts_file(
            db,
            file_name=file_name,
            file_bytes=payload,
            source_type=source_type,
            sheet_name=sheet_name,
            header_row=header_row,
        )
        result.warnings = [
            "Imported from Google Sheets URL. Private/authenticated sheets are not supported yet.",
            *result.warnings,
        ]
        return result
